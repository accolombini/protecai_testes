"""
================================================================================
RELAY CONFIGURATION REPORT SERVICE
================================================================================
Author: ProtecAI Engineering Team
Project: PETRO_ProtecAI
Date: 2025-11-03
Version: 1.0.0

Description:
    Serviço especializado para geração de relatórios de configuração (setup)
    de relés de proteção. Consolida dados de:
    - protec_ai.relay_equipment
    - protec_ai.protection_functions
    - protec_ai.relay_settings
    - protec_ai.vw_protection_settings (view otimizada)
    
    Formatos de saída: JSON, CSV, XLSX, PDF

Principles:
    - ZERO MOCK: Dados reais do banco
    - CAUSA RAIZ: Queries consolidadas, não hardcoded
    - ROBUSTEZ: Tratamento de erros completo
    - RASTREABILIDADE: Logs e auditoria

Usage:
    from api.services.relay_config_report_service import RelayConfigReportService
    
    service = RelayConfigReportService(db_session)
    report = service.generate_configuration_report(equipment_id=1, format='json')
================================================================================
"""

import logging
from typing import Dict, List, Any, Optional
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import text
from fastapi import HTTPException
import csv
import io
import json

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    logging.warning("⚠️  reportlab não instalado. Exportação PDF desabilitada.")

# Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logging.warning("⚠️  openpyxl não instalado. Exportação XLSX desabilitada.")

logger = logging.getLogger(__name__)


class RelayConfigReportService:
    """
    Serviço de geração de relatórios de configuração de relés.
    
    Attributes:
        db: SQLAlchemy session
    """
    
    def __init__(self, db: Session):
        """
        Inicializa o serviço.
        
        Args:
            db: Sessão SQLAlchemy
        """
        self.db = db
    
    def get_equipment_info(self, equipment_id: int) -> Optional[Dict[str, Any]]:
        """
        Obtém informações básicas do equipamento.
        
        Args:
            equipment_id: ID do equipamento
        
        Returns:
            Dicionário com dados do equipamento ou None
        """
        try:
            query = text("""
                SELECT 
                    e.id,
                    e.equipment_tag,
                    e.plant_reference,
                    e.bay_position,
                    e.serial_number,
                    e.software_version,
                    e.description,
                    e.frequency,
                    e.status,
                    m.name as model_name,
                    m.model_type,
                    m.voltage_class,
                    mf.name as manufacturer_name,
                    ec.phase_ct_primary,
                    ec.phase_ct_secondary,
                    ec.neutral_ct_primary,
                    ec.neutral_ct_secondary,
                    ec.vt_primary,
                    ec.vt_secondary
                FROM protec_ai.relay_equipment e
                LEFT JOIN protec_ai.relay_models m ON e.model_id = m.id
                LEFT JOIN protec_ai.manufacturers mf ON m.manufacturer_id = mf.id
                LEFT JOIN protec_ai.electrical_configuration ec ON e.id = ec.equipment_id
                WHERE e.id = :equipment_id
            """)
            
            result = self.db.execute(query, {"equipment_id": equipment_id}).fetchone()
            
            if not result:
                return None
            
            return dict(result._mapping)
            
        except Exception as e:
            logger.error(f"❌ Erro ao obter info do equipamento {equipment_id}: {e}")
            return None
    
    def get_protection_functions(self, equipment_id: int) -> List[Dict[str, Any]]:
        """
        Obtém funções de proteção configuradas no equipamento.
        
        Args:
            equipment_id: ID do equipamento
        
        Returns:
            Lista de funções de proteção
        """
        try:
            query = text("""
                SELECT 
                    pf.id,
                    pf.function_code,
                    pf.function_name,
                    pf.function_description,
                    pf.ansi_ieee_standard,
                    pf.is_primary,
                    epf.is_enabled,
                    epf.priority
                FROM protec_ai.protection_functions pf
                LEFT JOIN protec_ai.equipment_protection_functions epf 
                    ON pf.id = epf.function_id 
                    AND epf.equipment_id = :equipment_id
                WHERE epf.equipment_id = :equipment_id
                ORDER BY pf.function_code, pf.function_name
            """)
            
            results = self.db.execute(query, {"equipment_id": equipment_id}).fetchall()
            
            return [dict(row._mapping) for row in results]
            
        except Exception as e:
            logger.error(f"❌ Erro ao obter funções de proteção: {e}")
            return []
    
    def get_relay_settings(self, equipment_id: int, function_id: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Obtém configurações/ajustes do relé.
        
        Args:
            equipment_id: ID do equipamento
            function_id: ID da função (opcional, para filtrar)
        
        Returns:
            Lista de configurações
        """
        try:
            if function_id:
                query = text("""
                    SELECT 
                        rs.id,
                        rs.parameter_name,
                        rs.parameter_code,
                        rs.set_value,
                        rs.set_value_text,
                        rs.unit_of_measure,
                        rs.setting_group,
                        rs.is_enabled,
                        pf.function_code,
                        pf.function_name
                    FROM protec_ai.relay_settings rs
                    LEFT JOIN protec_ai.protection_functions pf ON rs.function_id = pf.id
                    WHERE rs.equipment_id = :equipment_id
                      AND rs.function_id = :function_id
                    ORDER BY rs.parameter_code
                """)
                params = {"equipment_id": equipment_id, "function_id": function_id}
            else:
                query = text("""
                    SELECT 
                        rs.id,
                        rs.parameter_name,
                        rs.parameter_code,
                        rs.set_value,
                        rs.set_value_text,
                        rs.unit_of_measure,
                        rs.setting_group,
                        rs.is_enabled,
                        pf.function_code,
                        pf.function_name
                    FROM protec_ai.relay_settings rs
                    LEFT JOIN protec_ai.protection_functions pf ON rs.function_id = pf.id
                    WHERE rs.equipment_id = :equipment_id
                    ORDER BY pf.function_code, rs.parameter_code
                """)
                params = {"equipment_id": equipment_id}
            
            results = self.db.execute(query, params).fetchall()
            
            return [dict(row._mapping) for row in results]
            
        except Exception as e:
            logger.error(f"❌ Erro ao obter configurações do relé: {e}")
            return []
    
    def generate_configuration_report(
        self,
        equipment_id: int,
        format: str = 'json',
        include_disabled: bool = False
    ) -> Dict[str, Any]:
        """
        Gera relatório consolidado de configuração do equipamento.
        
        Args:
            equipment_id: ID do equipamento
            format: Formato de saída ('json', 'csv', 'xlsx', 'pdf')
            include_disabled: Incluir funções/parâmetros desabilitados
        
        Returns:
            Dicionário com relatório e metadados
        
        Raises:
            HTTPException: Se equipamento não encontrado ou erro de geração
        """
        try:
            logger.info(f"📋 Gerando relatório de configuração: equipment_id={equipment_id}, format={format}")
            
            # 1. Obter informações do equipamento
            equipment = self.get_equipment_info(equipment_id)
            if not equipment:
                raise HTTPException(
                    status_code=404,
                    detail=f"Equipamento ID {equipment_id} não encontrado"
                )
            
            # 2. Obter funções de proteção
            functions = self.get_protection_functions(equipment_id)
            
            # Filtrar desabilitadas se necessário
            if not include_disabled:
                functions = [f for f in functions if f.get('is_enabled', False)]
            
            # 3. Obter configurações para cada função
            functions_with_settings = []
            for func in functions:
                settings = self.get_relay_settings(equipment_id, func['id'])
                
                # Filtrar desabilitadas se necessário
                if not include_disabled:
                    settings = [s for s in settings if s.get('is_enabled', True)]
                
                functions_with_settings.append({
                    **func,
                    'settings': settings,
                    'settings_count': len(settings)
                })
            
            # 4. Obter todas as configurações gerais (sem função específica)
            general_settings = self.get_relay_settings(equipment_id)
            # Remover as que já estão associadas a funções
            function_setting_ids = set()
            for func in functions_with_settings:
                function_setting_ids.update([s['id'] for s in func['settings']])
            
            general_settings = [
                s for s in general_settings 
                if s['id'] not in function_setting_ids
            ]
            
            # 5. Montar relatório consolidado
            report = {
                'equipment': equipment,
                'protection_functions': functions_with_settings,
                'general_settings': general_settings,
                'summary': {
                    'total_functions': len(functions_with_settings),
                    'enabled_functions': len([f for f in functions_with_settings if f.get('is_enabled', False)]),
                    'total_settings': sum(f['settings_count'] for f in functions_with_settings) + len(general_settings),
                    'report_generated_at': datetime.now().isoformat(),
                    'include_disabled': include_disabled
                }
            }
            
            # 6. Gerar formato solicitado
            if format.lower() == 'json':
                return {'data': report, 'format': 'json'}
            
            elif format.lower() == 'csv':
                csv_content = self._generate_csv(report)
                return {'data': csv_content, 'format': 'csv', 'filename': self._get_filename(equipment, 'csv')}
            
            elif format.lower() == 'xlsx':
                if not HAS_OPENPYXL:
                    raise HTTPException(status_code=500, detail="openpyxl não instalado")
                xlsx_content = self._generate_xlsx(report)
                return {'data': xlsx_content, 'format': 'xlsx', 'filename': self._get_filename(equipment, 'xlsx')}
            
            elif format.lower() == 'pdf':
                if not HAS_REPORTLAB:
                    raise HTTPException(status_code=500, detail="reportlab não instalado")
                pdf_content = self._generate_pdf(report)
                return {'data': pdf_content, 'format': 'pdf', 'filename': self._get_filename(equipment, 'pdf')}
            
            else:
                raise HTTPException(status_code=400, detail=f"Formato inválido: {format}")
        
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"❌ Erro ao gerar relatório: {e}")
            raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")
    
    def _get_filename(self, equipment: Dict[str, Any], extension: str) -> str:
        """Gera nome de arquivo para download."""
        tag = equipment.get('equipment_tag', 'UNKNOWN').replace('/', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"CONFIG_{tag}_{timestamp}.{extension}"
    
    def _generate_csv(self, report: Dict[str, Any]) -> str:
        """Gera CSV do relatório."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header do equipamento
        equipment = report['equipment']
        writer.writerow(['RELATÓRIO DE CONFIGURAÇÃO DE RELÉ'])
        writer.writerow([])
        writer.writerow(['Tag', equipment.get('equipment_tag', '')])
        writer.writerow(['Fabricante', equipment.get('manufacturer_name', '')])
        writer.writerow(['Modelo', equipment.get('model_name', '')])
        writer.writerow(['Serial', equipment.get('serial_number', '')])
        writer.writerow([])
        
        # Funções de proteção e settings
        writer.writerow(['FUNÇÕES DE PROTEÇÃO E CONFIGURAÇÕES'])
        writer.writerow([])
        
        for func in report['protection_functions']:
            writer.writerow([])
            writer.writerow([f"Função: {func['function_code']} - {func['function_name']}", '', f"Habilitada: {func.get('is_enabled', False)}"])
            writer.writerow(['Parâmetro', 'Código', 'Valor', 'Unidade', 'Grupo'])
            
            for setting in func['settings']:
                value = setting.get('set_value_text') or setting.get('set_value') or ''
                writer.writerow([
                    setting.get('parameter_name', ''),
                    setting.get('parameter_code', ''),
                    value,
                    setting.get('unit_of_measure', ''),
                    setting.get('setting_group', '')
                ])
        
        # Configurações gerais
        if report['general_settings']:
            writer.writerow([])
            writer.writerow(['CONFIGURAÇÕES GERAIS'])
            writer.writerow(['Parâmetro', 'Código', 'Valor', 'Unidade', 'Grupo'])
            for setting in report['general_settings']:
                value = setting.get('set_value_text') or setting.get('set_value') or ''
                writer.writerow([
                    setting.get('parameter_name', ''),
                    setting.get('parameter_code', ''),
                    value,
                    setting.get('unit_of_measure', ''),
                    setting.get('setting_group', '')
                ])
        
        return output.getvalue()
    
    def _generate_xlsx(self, report: Dict[str, Any]) -> bytes:
        """Gera XLSX do relatório."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Configuração"
        
        # Estilos
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        table_header_font = Font(bold=True, color="FFFFFF")
        
        row = 1
        equipment = report['equipment']
        
        # Header
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "RELATÓRIO DE CONFIGURAÇÃO DE RELÉ"
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        row += 2
        
        # Info do equipamento
        ws[f'A{row}'] = "Tag:"
        ws[f'B{row}'] = equipment.get('equipment_tag', '')
        row += 1
        ws[f'A{row}'] = "Fabricante:"
        ws[f'B{row}'] = equipment.get('manufacturer_name', '')
        row += 1
        ws[f'A{row}'] = "Modelo:"
        ws[f'B{row}'] = equipment.get('model_name', '')
        row += 1
        ws[f'A{row}'] = "Serial:"
        ws[f'B{row}'] = equipment.get('serial_number', '')
        row += 2
        
        # Funções e settings
        for func in report['protection_functions']:
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = f"{func['function_code']} - {func['function_name']} (Habilitada: {func.get('is_enabled', False)})"
            cell.font = subheader_font
            row += 1
            
            # Header da tabela
            headers = ['Parâmetro', 'Código', 'Valor', 'Unidade', 'Grupo']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = table_header_fill
                cell.font = table_header_font
            row += 1
            
            # Dados
            for setting in func['settings']:
                value = setting.get('set_value_text') or setting.get('set_value') or ''
                ws.cell(row=row, column=1, value=setting.get('parameter_name', ''))
                ws.cell(row=row, column=2, value=setting.get('parameter_code', ''))
                ws.cell(row=row, column=3, value=str(value))
                ws.cell(row=row, column=4, value=setting.get('unit_of_measure', ''))
                ws.cell(row=row, column=5, value=setting.get('setting_group', ''))
                row += 1
            row += 1
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        
        # Salvar em bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _generate_pdf(self, report: Dict[str, Any]) -> bytes:
        """Gera PDF do relatório."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f497d'),
            alignment=TA_CENTER
        )
        
        equipment = report['equipment']
        elements.append(Paragraph("RELATÓRIO DE CONFIGURAÇÃO DE RELÉ", title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Info do equipamento
        info_data = [
            ['Tag:', equipment.get('equipment_tag', '')],
            ['Fabricante:', equipment.get('manufacturer_name', '')],
            ['Modelo:', equipment.get('model_name', '')],
            ['Serial:', equipment.get('serial_number', '')],
        ]
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Funções e settings
        for func in report['protection_functions']:
            # Título da função
            func_title = Paragraph(
                f"<b>{func['function_code']} - {func['function_name']}</b>",
                styles['Heading2']
            )
            elements.append(func_title)
            elements.append(Spacer(1, 0.1*inch))
            
            # Tabela de settings
            table_data = [['Parâmetro', 'Código', 'Valor', 'Unidade']]
            for setting in func['settings']:
                value = setting.get('set_value_text') or setting.get('set_value') or ''
                table_data.append([
                    setting.get('parameter_name', ''),
                    setting.get('parameter_code', ''),
                    str(value),
                    setting.get('unit_of_measure', '')
                ])
            
            t = Table(table_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 0.8*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 0.2*inch))
        
        doc.build(elements)
        return buffer.getvalue()
