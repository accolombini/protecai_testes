#!/usr/bin/env python3
"""
🔬 ANÁLISE UNIVERSAL DO GLOSSÁRIO - EXTRAÇÃO COMPLETA
=======================================================

Analisa profundamente o glossário para criar um parser universal que:
1. Identifica estrutura hierárquica (seções/subseções)
2. Detecta avisos e condicionais (células amarelas)
3. Mapeia relações entre parâmetros (G1 → G2, referências)
4. Extrai metadados de formatação (cores, estilos)
5. Identifica padrões de nomenclatura universais

Objetivo: Criar estrutura que permita parsing genérico sem lógica específica por modelo.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
import json
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, asdict
from collections import defaultdict
import re

@dataclass
class CellMetadata:
    """Metadados completos de uma célula"""
    row: int
    col: int
    value: Any
    bg_color: Optional[str] = None
    font_color: Optional[str] = None
    is_bold: bool = False
    is_italic: bool = False
    font_size: Optional[int] = None
    is_merged: bool = False
    
@dataclass
class Section:
    """Representa uma seção do glossário"""
    name: str
    level: int  # 0=raiz, 1=seção, 2=subseção
    start_row: int
    end_row: Optional[int] = None
    parent: Optional[str] = None
    metadata: Optional[CellMetadata] = None
    warnings: List[str] = None
    conditionals: List[str] = None
    
    def __post_init__(self):
        if self.warnings is None:
            self.warnings = []
        if self.conditionals is None:
            self.conditionals = []

@dataclass
class Parameter:
    """Representa um parâmetro de relé"""
    code: str
    name: str
    value: Any
    section: str
    subsection: Optional[str] = None
    unit: Optional[str] = None
    options: List[str] = None
    conditions: List[str] = None
    row: int = 0
    col: int = 0
    
    def __post_init__(self):
        if self.options is None:
            self.options = []
        if self.conditions is None:
            self.conditions = []

class UniversalGlossarioAnalyzer:
    """Analisador universal de glossários de relés"""
    
    # Cores conhecidas (RGB em hexadecimal)
    YELLOW_VARIANTS = ['FFFFFF00', 'FFFF00', 'FFFFE0', 'FFFFCC', 'FFFF99', 'FFFACD']
    
    # Padrões de código de parâmetro
    PARAM_CODE_PATTERNS = [
        r'^[0-9A-F]{4}:',  # 0104:, 02A0:, etc
        r'^[0-9]{2}\.[0-9]{2}[A-Z]?:',  # 00.04:, 42.0D:, etc
    ]
    
    # Palavras-chave de seção
    SECTION_KEYWORDS = [
        'PARAMETERS', 'CONFIGURATION', 'PROTECTION', 'AUTOMAT', 'SYSTEM DATA',
        'CT AND VT', 'CB MONITOR', 'Group', 'Caracteristiques'
    ]
    
    # Palavras-chave de aviso/condicional
    WARNING_KEYWORDS = [
        'Repete', 'Idem', 'caso', 'mesmos dados', 'execeção', 'introduzidos',
        'apenas se', 'somente quando', 'se estiver'
    ]
    
    def __init__(self, glossario_path: str):
        self.glossario_path = Path(glossario_path)
        self.workbook = openpyxl.load_workbook(glossario_path, data_only=False)
        self.analysis_results = {}
        
    def analyze_all_sheets(self) -> Dict[str, Any]:
        """Analisa todas as abas do glossário"""
        print("=" * 80)
        print("🔬 ANÁLISE UNIVERSAL DO GLOSSÁRIO")
        print("=" * 80)
        print(f"\n📁 Arquivo: {self.glossario_path.name}")
        print(f"📊 Total de abas: {len(self.workbook.sheetnames)}\n")
        
        all_results = {
            'file': str(self.glossario_path),
            'total_sheets': len(self.workbook.sheetnames),
            'sheets': {},
            'universal_patterns': {
                'section_hierarchy': {},
                'param_code_formats': set(),
                'warning_types': defaultdict(int),
                'conditional_patterns': set(),
                'value_formats': defaultdict(int)
            }
        }
        
        for sheet_name in self.workbook.sheetnames:
            print(f"\n{'=' * 80}")
            print(f"📋 ANALISANDO ABA: {sheet_name}")
            print("=" * 80)
            
            result = self.analyze_sheet(sheet_name)
            all_results['sheets'][sheet_name] = result
            
            # Acumular padrões universais
            self._accumulate_universal_patterns(result, all_results['universal_patterns'])
            
        # Converter sets para listas para JSON
        all_results['universal_patterns']['param_code_formats'] = list(
            all_results['universal_patterns']['param_code_formats']
        )
        all_results['universal_patterns']['conditional_patterns'] = list(
            all_results['universal_patterns']['conditional_patterns']
        )
        
        self._print_universal_summary(all_results)
        return all_results
    
    def analyze_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """Analisa uma aba específica"""
        ws = self.workbook[sheet_name]
        
        # Extrair metadados de todas as células
        cells_metadata = self._extract_all_cells_metadata(ws)
        
        # Identificar seções (células amarelas)
        sections = self._identify_sections(cells_metadata)
        
        # Extrair parâmetros
        parameters = self._extract_parameters(ws, sections, cells_metadata)
        
        # Identificar avisos e condicionais
        warnings, conditionals = self._extract_warnings_and_conditionals(cells_metadata)
        
        # Detectar estrutura hierárquica
        hierarchy = self._build_hierarchy(sections)
        
        # Identificar modelo/fabricante
        metadata = self._extract_sheet_metadata(ws)
        
        result = {
            'metadata': metadata,
            'sections': [asdict(s) for s in sections],
            'hierarchy': hierarchy,
            'parameters': [asdict(p) for p in parameters],
            'warnings': warnings,
            'conditionals': conditionals,
            'statistics': {
                'total_sections': len(sections),
                'total_parameters': len(parameters),
                'total_warnings': len(warnings),
                'total_conditionals': len(conditionals)
            }
        }
        
        self._print_sheet_summary(result)
        return result
    
    def _extract_all_cells_metadata(self, ws) -> Dict[Tuple[int, int], CellMetadata]:
        """Extrai metadados de todas as células"""
        cells = {}
        
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is None:
                    continue
                    
                metadata = CellMetadata(
                    row=row_idx,
                    col=col_idx,
                    value=cell.value,
                    bg_color=self._get_bg_color(cell),
                    font_color=self._get_font_color(cell),
                    is_bold=self._is_bold(cell),
                    is_italic=self._is_italic(cell),
                    font_size=self._get_font_size(cell),
                    is_merged=cell.coordinate in ws.merged_cells
                )
                
                cells[(row_idx, col_idx)] = metadata
        
        return cells
    
    def _get_bg_color(self, cell) -> Optional[str]:
        """Obtém cor de fundo da célula"""
        if cell.fill and cell.fill.fgColor:
            color = cell.fill.fgColor.rgb
            if color and isinstance(color, str):
                return color
        return None
    
    def _get_font_color(self, cell) -> Optional[str]:
        """Obtém cor da fonte"""
        if cell.font and cell.font.color:
            color = cell.font.color.rgb
            if color and isinstance(color, str):
                return color
        return None
    
    def _is_bold(self, cell) -> bool:
        """Verifica se célula está em negrito"""
        return cell.font and cell.font.bold is True
    
    def _is_italic(self, cell) -> bool:
        """Verifica se célula está em itálico"""
        return cell.font and cell.font.italic is True
    
    def _get_font_size(self, cell) -> Optional[int]:
        """Obtém tamanho da fonte"""
        if cell.font and cell.font.size:
            return int(cell.font.size)
        return None
    
    def _is_yellow_cell(self, metadata: CellMetadata) -> bool:
        """Verifica se célula tem fundo amarelo"""
        if not metadata.bg_color:
            return False
        return any(yellow in metadata.bg_color.upper() for yellow in self.YELLOW_VARIANTS)
    
    def _identify_sections(self, cells_metadata: Dict) -> List[Section]:
        """Identifica seções baseadas em células amarelas e palavras-chave"""
        sections = []
        
        for (row, col), metadata in sorted(cells_metadata.items()):
            value_str = str(metadata.value).strip()
            
            # Células amarelas são sempre seções
            if self._is_yellow_cell(metadata):
                section = Section(
                    name=value_str,
                    level=self._determine_section_level(value_str, metadata),
                    start_row=row,
                    metadata=metadata
                )
                
                # Verificar se contém avisos/condicionais
                if any(kw in value_str for kw in self.WARNING_KEYWORDS):
                    section.warnings.append(value_str)
                
                sections.append(section)
            
            # Também considerar células com palavras-chave de seção
            elif any(kw in value_str for kw in self.SECTION_KEYWORDS):
                if metadata.is_bold or metadata.font_size and metadata.font_size > 11:
                    section = Section(
                        name=value_str,
                        level=self._determine_section_level(value_str, metadata),
                        start_row=row,
                        metadata=metadata
                    )
                    sections.append(section)
        
        # Definir end_row de cada seção
        for i, section in enumerate(sections):
            if i < len(sections) - 1:
                section.end_row = sections[i + 1].start_row - 1
        
        return sections
    
    def _determine_section_level(self, name: str, metadata: CellMetadata) -> int:
        """Determina nível hierárquico da seção"""
        # Nível 0: Títulos principais (OP PARAMETERS, PROTECTION, etc)
        if any(kw in name.upper() for kw in ['PARAMETERS', 'PROTECTION', 'CONFIGURATION', 'SYSTEM DATA']):
            return 0
        
        # Nível 1: Subseções (Group 1, Group 2, funções específicas)
        if 'GROUP' in name.upper() or 'G1' in name.upper() or 'G2' in name.upper():
            return 1
        
        # Nível 2: Funções específicas ([50N/51N], [59], etc)
        if re.search(r'\[\d+[A-Z]*\]', name):
            return 2
        
        # Baseado na coluna (quanto mais à direita, mais específico)
        if metadata.col <= 3:
            return 0
        elif metadata.col <= 5:
            return 1
        else:
            return 2
    
    def _extract_parameters(self, ws, sections: List[Section], 
                           cells_metadata: Dict) -> List[Parameter]:
        """Extrai parâmetros de configuração"""
        parameters = []
        
        for (row, col), metadata in sorted(cells_metadata.items()):
            value_str = str(metadata.value).strip()
            
            # Verificar se é um código de parâmetro
            for pattern in self.PARAM_CODE_PATTERNS:
                if re.match(pattern, value_str):
                    param = self._parse_parameter(value_str, row, col, sections, ws)
                    if param:
                        parameters.append(param)
                    break
        
        return parameters
    
    def _parse_parameter(self, text: str, row: int, col: int, 
                        sections: List[Section], ws) -> Optional[Parameter]:
        """Faz parsing de um parâmetro"""
        # Extrair código e nome
        parts = text.split(':', 1)
        if len(parts) < 2:
            return None
        
        code = parts[0].strip()
        rest = parts[1].strip()
        
        # Tentar encontrar nome e valor
        # Padrão: "Nome: Valor" ou "Nome :: Valor"
        if '::' in rest:
            name, value = rest.split('::', 1)
        elif ':' in rest:
            name, value = rest.split(':', 1)
        else:
            name = rest
            value = ws.cell(row, col + 1).value if col + 1 <= ws.max_column else None
        
        name = name.strip()
        value = str(value).strip() if value else None
        
        # Extrair unidade
        unit = self._extract_unit(value) if value else None
        
        # Extrair opções (YES/NO, valores separados por /)
        options = self._extract_options(value) if value else []
        
        # Determinar seção
        section_name = self._find_section_for_row(row, sections)
        
        return Parameter(
            code=code,
            name=name,
            value=value,
            section=section_name,
            unit=unit,
            options=options,
            row=row,
            col=col
        )
    
    def _extract_unit(self, value: str) -> Optional[str]:
        """Extrai unidade de medida do valor"""
        # Padrões comuns: Hz, In, s, A, V, kV, etc
        units = ['Hz', 'In', 'kV', 'kA', 's', 'ms', 'A', 'V', 'MW', 'MVA', '°', 'deg']
        for unit in units:
            if unit in value:
                return unit
        return None
    
    def _extract_options(self, value: str) -> List[str]:
        """Extrai opções possíveis do valor"""
        options = []
        
        # Padrão "YES / NO"
        if '/' in value:
            options = [opt.strip() for opt in value.split('/')]
        
        # Padrão "YES YES / NO"
        elif ' YES ' in value.upper() and ' NO' in value.upper():
            options = ['YES', 'NO']
        
        return options
    
    def _find_section_for_row(self, row: int, sections: List[Section]) -> str:
        """Encontra seção que contém a linha"""
        for section in reversed(sections):  # Buscar de trás para frente
            if section.start_row <= row:
                if section.end_row is None or row <= section.end_row:
                    return section.name
        return "Unknown"
    
    def _extract_warnings_and_conditionals(self, cells_metadata: Dict) -> Tuple[List[str], List[str]]:
        """Extrai avisos e condicionais"""
        warnings = []
        conditionals = []
        
        for metadata in cells_metadata.values():
            value_str = str(metadata.value).strip()
            
            # Verificar avisos
            if any(kw in value_str for kw in ['Repete', 'Idem', 'mesmos dados', 'execeção']):
                warnings.append({
                    'row': metadata.row,
                    'col': metadata.col,
                    'text': value_str,
                    'is_yellow': self._is_yellow_cell(metadata)
                })
            
            # Verificar condicionais
            if any(kw in value_str for kw in ['caso', 'se', 'apenas', 'somente quando']):
                conditionals.append({
                    'row': metadata.row,
                    'col': metadata.col,
                    'text': value_str,
                    'is_yellow': self._is_yellow_cell(metadata)
                })
        
        return warnings, conditionals
    
    def _build_hierarchy(self, sections: List[Section]) -> Dict[str, Any]:
        """Constrói hierarquia de seções"""
        hierarchy = {
            'root': [],
            'relationships': []
        }
        
        # Ordenar por nível
        level_0 = [s for s in sections if s.level == 0]
        level_1 = [s for s in sections if s.level == 1]
        level_2 = [s for s in sections if s.level == 2]
        
        # Construir relacionamentos
        for parent in level_0:
            children = [s for s in level_1 
                       if s.start_row > parent.start_row and 
                       (parent.end_row is None or s.start_row <= parent.end_row)]
            
            if children:
                hierarchy['relationships'].append({
                    'parent': parent.name,
                    'children': [c.name for c in children]
                })
        
        hierarchy['root'] = [s.name for s in level_0]
        
        return hierarchy
    
    def _extract_sheet_metadata(self, ws) -> Dict[str, Any]:
        """Extrai metadados da aba (fabricante, modelo, tipo)"""
        metadata = {
            'manufacturer': None,
            'model': None,
            'type': None,
            'notes': []
        }
        
        # Procurar nas primeiras 10 linhas
        for row in ws.iter_rows(max_row=10):
            for cell in row:
                if cell.value:
                    value_str = str(cell.value).strip()
                    
                    if 'Fabricante' in value_str:
                        # Extrair fabricante
                        if ':' in value_str:
                            parts = value_str.split(':', 1)
                            if len(parts) > 1:
                                metadata['manufacturer'] = parts[1].strip().split('-')[0].strip()
                                if '-' in parts[1]:
                                    metadata['model'] = parts[1].strip().split('-', 1)[1].strip()
                    
                    if '(Relé' in value_str:
                        # Extrair tipo
                        match = re.search(r'\((.*?)\)', value_str)
                        if match:
                            metadata['type'] = match.group(1)
        
        return metadata
    
    def _accumulate_universal_patterns(self, sheet_result: Dict, universal: Dict):
        """Acumula padrões universais de todas as abas"""
        # Formatos de código
        for param in sheet_result['parameters']:
            code = param['code']
            for pattern in self.PARAM_CODE_PATTERNS:
                if re.match(pattern, code + ':'):
                    universal['param_code_formats'].add(pattern)
        
        # Tipos de aviso
        for warning in sheet_result['warnings']:
            text = warning['text']
            for keyword in self.WARNING_KEYWORDS:
                if keyword in text:
                    universal['warning_types'][keyword] += 1
        
        # Padrões condicionais
        for cond in sheet_result['conditionals']:
            universal['conditional_patterns'].add(cond['text'])
    
    def _print_sheet_summary(self, result: Dict):
        """Imprime resumo da análise da aba"""
        meta = result['metadata']
        stats = result['statistics']
        
        print(f"\n📌 FABRICANTE: {meta.get('manufacturer', 'N/A')}")
        print(f"📌 MODELO: {meta.get('model', 'N/A')}")
        print(f"📌 TIPO: {meta.get('type', 'N/A')}")
        
        print(f"\n📊 ESTATÍSTICAS:")
        print(f"   • Seções identificadas: {stats['total_sections']}")
        print(f"   • Parâmetros extraídos: {stats['total_parameters']}")
        print(f"   • Avisos encontrados: {stats['total_warnings']}")
        print(f"   • Condicionais encontrados: {stats['total_conditionals']}")
        
        if result['warnings']:
            print(f"\n⚠️  AVISOS:")
            for i, warning in enumerate(result['warnings'][:5], 1):
                print(f"   {i}. [{warning['row']}, {warning['col']}] {warning['text'][:80]}")
        
        if result['conditionals']:
            print(f"\n🔀 CONDICIONAIS:")
            for i, cond in enumerate(result['conditionals'][:5], 1):
                print(f"   {i}. [{cond['row']}, {cond['col']}] {cond['text'][:80]}")
        
        print(f"\n🌲 HIERARQUIA:")
        for parent in result['hierarchy']['root']:
            print(f"   📁 {parent}")
    
    def _print_universal_summary(self, all_results: Dict):
        """Imprime resumo de padrões universais"""
        print("\n" + "=" * 80)
        print("🌍 PADRÕES UNIVERSAIS IDENTIFICADOS")
        print("=" * 80)
        
        patterns = all_results['universal_patterns']
        
        print(f"\n📝 FORMATOS DE CÓDIGO DE PARÂMETRO:")
        for fmt in patterns['param_code_formats']:
            print(f"   • {fmt}")
        
        print(f"\n⚠️  TIPOS DE AVISO (TOP 5):")
        for keyword, count in sorted(patterns['warning_types'].items(), 
                                     key=lambda x: x[1], reverse=True)[:5]:
            print(f"   • '{keyword}': {count} ocorrências")
        
        print(f"\n🔀 PADRÕES CONDICIONAIS ÚNICOS: {len(patterns['conditional_patterns'])}")
        for i, pattern in enumerate(list(patterns['conditional_patterns'])[:5], 1):
            print(f"   {i}. {pattern[:80]}")
    
    def save_results(self, results: Dict, output_path: str):
        """Salva resultados da análise"""
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n✅ Análise completa salva em: {output_file}")


def main():
    glossario_path = "inputs/glossario/Dados_Glossario_Micon_Sepam.xlsx"
    
    if not Path(glossario_path).exists():
        print(f"❌ Arquivo não encontrado: {glossario_path}")
        return
    
    analyzer = UniversalGlossarioAnalyzer(glossario_path)
    results = analyzer.analyze_all_sheets()
    
    # Salvar resultados
    analyzer.save_results(results, "outputs/logs/glossario_universal_analysis.json")
    
    print("\n" + "=" * 80)
    print("✅ ANÁLISE COMPLETA!")
    print("=" * 80)


if __name__ == "__main__":
    main()
