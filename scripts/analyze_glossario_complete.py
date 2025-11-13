#!/usr/bin/env python3
"""
Análise COMPLETA do glossário Excel - todas as abas com observações e instruções.
Objetivo: Entender estrutura exata para extração correta de parâmetros dos relés.
"""

import openpyxl
import pandas as pd
from pathlib import Path
import json

# Caminhos
GLOSSARIO_PATH = Path("/Users/accol/Library/Mobile Documents/com~apple~CloudDocs/UNIVERSIDADES/UFF/PROJETOS/PETROBRAS/PETRO_ProtecAI/protecai_testes/inputs/glossario/Dados_Glossario_Micon_Sepam.xlsx")

def analyze_excel_structure():
    """Analisa estrutura completa do Excel incluindo todas as abas e observações."""
    
    print("=" * 80)
    print("📋 ANÁLISE COMPLETA DO GLOSSÁRIO - TODAS AS ABAS")
    print("=" * 80)
    
    # Carregar workbook com openpyxl (preserva formatação e comentários)
    wb = openpyxl.load_workbook(GLOSSARIO_PATH, data_only=False)
    
    print(f"\n📁 Arquivo: {GLOSSARIO_PATH.name}")
    print(f"📊 Total de abas: {len(wb.sheetnames)}\n")
    
    analysis_results = {}
    
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"📄 ABA: {sheet_name}")
        print(f"{'='*80}")
        
        sheet = wb[sheet_name]
        
        # Informações básicas da aba
        print(f"Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas")
        
        # Ler com pandas para análise de dados
        df = pd.read_excel(GLOSSARIO_PATH, sheet_name=sheet_name)
        
        print(f"\nColunas encontradas ({len(df.columns)}):")
        for i, col in enumerate(df.columns, 1):
            print(f"  {i}. {col}")
        
        # Mostrar primeiras linhas
        print(f"\n📝 PRIMEIRAS 10 LINHAS DE DADOS:")
        print(df.head(10).to_string(max_colwidth=50))
        
        # Procurar observações/instruções (células mescladas ou com formatação especial)
        print(f"\n🔍 PROCURANDO OBSERVAÇÕES E INSTRUÇÕES ESPECIAIS:")
        
        observations = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(50, sheet.max_row)), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Detectar células com instruções (keywords)
                    keywords = ['obs:', 'observação', 'atenção', 'nota:', 'importante', 
                               'instrução', 'regra', 'considerar', 'extrair', 'calcular']
                    
                    cell_text = str(cell.value).lower()
                    if any(keyword in cell_text for keyword in keywords):
                        observations.append({
                            'linha': row_idx,
                            'coluna': cell.column_letter,
                            'valor': cell.value[:200]  # Limitar tamanho
                        })
        
        if observations:
            print(f"  ✅ {len(observations)} observações/instruções encontradas:")
            for obs in observations[:10]:  # Mostrar até 10
                print(f"    • Linha {obs['linha']}, Coluna {obs['coluna']}:")
                print(f"      {obs['valor']}")
        else:
            print("  ℹ️  Nenhuma observação explícita encontrada nos primeiros 50 linhas")
        
        # Estatísticas da aba
        print(f"\n📊 ESTATÍSTICAS:")
        print(f"  • Total de registros: {len(df)}")
        print(f"  • Registros com valores: {df.dropna(how='all').shape[0]}")
        print(f"  • Colunas com dados completos: {df.notna().all().sum()}")
        
        # Valores únicos em colunas-chave (se existirem)
        key_columns = ['Grupo', 'Função', 'Tipo', 'Categoria', 'Unidade', 'Parâmetro']
        for col in key_columns:
            if col in df.columns:
                unique_count = df[col].nunique()
                print(f"  • Valores únicos em '{col}': {unique_count}")
                if unique_count <= 10:
                    print(f"    Valores: {df[col].dropna().unique().tolist()}")
        
        # Armazenar resultado
        analysis_results[sheet_name] = {
            'dimensions': (sheet.max_row, sheet.max_column),
            'columns': df.columns.tolist(),
            'row_count': len(df),
            'observations_found': len(observations),
            'observations': observations[:5]  # Top 5
        }
    
    # Resumo final
    print("\n" + "=" * 80)
    print("📋 RESUMO GERAL DO GLOSSÁRIO")
    print("=" * 80)
    
    for sheet_name, info in analysis_results.items():
        print(f"\n{sheet_name}:")
        print(f"  • Dimensões: {info['dimensions'][0]} linhas x {info['dimensions'][1]} colunas")
        print(f"  • Colunas: {', '.join(info['columns'][:5])}...")
        print(f"  • Registros: {info['row_count']}")
        print(f"  • Observações: {info['observations_found']}")
    
    # Salvar análise completa em JSON
    output_path = Path("outputs/logs/glossario_analysis_complete.json")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(analysis_results, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Análise completa salva em: {output_path}")
    
    return analysis_results

if __name__ == "__main__":
    analyze_excel_structure()
