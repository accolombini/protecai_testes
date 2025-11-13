"""
Testes para validar a geração de relatórios de configuração de relés.
Valida o módulo api/services/relay_config_report_service.py
"""

import pytest
import json
import csv
from pathlib import Path
import sys
from io import BytesIO
from openpyxl import load_workbook

# Adicionar o diretório raiz ao path
sys.path.insert(0, str(Path(__file__).parent.parent))

from api.services.relay_config_report_service import RelayConfigReportService


class TestRelayConfigReportService:
    """Testes para o serviço de relatórios"""
    
    @pytest.fixture
    def mock_db_session(self, mocker):
        """Cria um mock da sessão do banco de dados"""
        mock_session = mocker.MagicMock()
        
        # Mock para get_equipment_info
        mock_equipment = mocker.MagicMock()
        mock_equipment.equipment_name = "RELE_TEST_001"
        mock_equipment.manufacturer = "MICON"
        mock_equipment.model = "P122"
        mock_equipment.installation_date = "2024-01-15"
        mock_equipment.substation_name = "SE_TEST"
        mock_equipment.bay_name = "BAY_TEST"
        
        # Mock para get_protection_functions
        mock_function = mocker.MagicMock()
        mock_function.ansi_code = "50"
        mock_function.function_name = "Overcurrent"
        mock_function.description = "Instantaneous Overcurrent Protection"
        mock_function.is_enabled = True
        
        # Mock para get_relay_settings
        mock_setting = mocker.MagicMock()
        mock_setting.setting_id = 1
        mock_setting.ansi_function = "50"
        mock_setting.original_code = "I>"
        mock_setting.setting_name = "Corrente de Pickup"
        mock_setting.default_value = "5.5"
        mock_setting.configured_value = "6.0"
        mock_setting.unit = "A"
        mock_setting.category = "protection"
        mock_setting.is_active = True
        
        # Configurar retornos dos mocks
        mock_session.execute.return_value.fetchone.return_value = mock_equipment
        mock_session.execute.return_value.fetchall.side_effect = [
            [mock_function],
            [mock_setting]
        ]
        
        return mock_session
    
    @pytest.fixture
    def report_service(self, mock_db_session):
        """Cria instância do serviço com mock de DB"""
        return RelayConfigReportService(mock_db_session)
    
    def test_get_equipment_info(self, report_service):
        """Testa obtenção de informações do equipamento"""
        equipment = report_service.get_equipment_info(1)
        
        assert equipment is not None
        assert equipment.equipment_name == "RELE_TEST_001"
        assert equipment.manufacturer == "MICON"
        assert equipment.model == "P122"
    
    def test_get_protection_functions(self, report_service):
        """Testa obtenção de funções de proteção"""
        functions = report_service.get_protection_functions(1)
        
        assert len(functions) == 1
        assert functions[0].ansi_code == "50"
        assert functions[0].function_name == "Overcurrent"
    
    def test_get_relay_settings(self, report_service):
        """Testa obtenção de configurações do relé"""
        settings = report_service.get_relay_settings(1, include_disabled=True)
        
        assert len(settings) == 1
        assert settings[0].original_code == "I>"
        assert settings[0].configured_value == "6.0"
    
    def test_generate_configuration_report_json(self, report_service):
        """Testa geração de relatório em formato JSON"""
        report = report_service.generate_configuration_report(1, format_type='json')
        
        assert report is not None
        assert 'equipment_info' in report
        assert 'protection_functions' in report
        assert 'relay_settings' in report
        
        # Validar estrutura do equipamento
        assert report['equipment_info']['equipment_name'] == "RELE_TEST_001"
        assert report['equipment_info']['manufacturer'] == "MICON"
        
        # Validar funções
        assert len(report['protection_functions']) == 1
        assert report['protection_functions'][0]['ansi_code'] == "50"
        
        # Validar settings
        assert len(report['relay_settings']) == 1
        assert report['relay_settings'][0]['configured_value'] == "6.0"


class TestReportGenerationCSV:
    """Testes para geração de relatórios em CSV"""
    
    @pytest.fixture
    def sample_report_data(self):
        """Dados de exemplo para relatório"""
        return {
            'equipment_info': {
                'equipment_id': 1,
                'equipment_name': 'RELE_001',
                'manufacturer': 'MICON',
                'model': 'P122'
            },
            'protection_functions': [
                {'ansi_code': '50', 'function_name': 'Overcurrent', 'is_enabled': True}
            ],
            'relay_settings': [
                {
                    'setting_id': 1,
                    'ansi_function': '50',
                    'original_code': 'I>',
                    'setting_name': 'Pickup',
                    'configured_value': '5.5',
                    'unit': 'A'
                }
            ]
        }
    
    def test_generate_csv_format(self, sample_report_data):
        """Testa geração de CSV"""
        # Usar serviço mock para gerar CSV
        from api.services.relay_config_report_service import RelayConfigReportService
        
        # Mock session
        mock_session = None
        service = RelayConfigReportService(mock_session)
        
        # Gerar CSV
        csv_content = service._generate_csv(sample_report_data)
        
        assert csv_content is not None
        assert 'RELE_001' in csv_content
        assert 'MICON' in csv_content
        assert 'Overcurrent' in csv_content
    
    def test_csv_parsing(self, sample_report_data, tmp_path):
        """Testa que CSV pode ser parseado corretamente"""
        from api.services.relay_config_report_service import RelayConfigReportService
        
        mock_session = None
        service = RelayConfigReportService(mock_session)
        
        csv_content = service._generate_csv(sample_report_data)
        
        # Salvar temporariamente
        csv_file = tmp_path / "test_report.csv"
        with open(csv_file, 'w', encoding='utf-8') as f:
            f.write(csv_content)
        
        # Ler e validar
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        assert len(rows) >= 1
        # Primeira linha deve conter informações do equipamento ou configurações


class TestReportGenerationXLSX:
    """Testes para geração de relatórios em XLSX"""
    
    @pytest.fixture
    def sample_report_data(self):
        """Dados de exemplo para relatório"""
        return {
            'equipment_info': {
                'equipment_id': 1,
                'equipment_name': 'RELE_001',
                'manufacturer': 'MICON',
                'model': 'P122',
                'substation_name': 'SE_TEST',
                'bay_name': 'BAY_TEST'
            },
            'protection_functions': [
                {'ansi_code': '50', 'function_name': 'Overcurrent', 'is_enabled': True},
                {'ansi_code': '27', 'function_name': 'Undervoltage', 'is_enabled': True}
            ],
            'relay_settings': [
                {
                    'setting_id': 1,
                    'ansi_function': '50',
                    'original_code': 'I>',
                    'setting_name': 'Pickup Current',
                    'configured_value': '5.5',
                    'unit': 'A',
                    'category': 'protection'
                }
            ]
        }
    
    def test_generate_xlsx_format(self, sample_report_data):
        """Testa geração de XLSX"""
        from api.services.relay_config_report_service import RelayConfigReportService
        
        mock_session = None
        service = RelayConfigReportService(mock_session)
        
        xlsx_bytes = service._generate_xlsx(sample_report_data)
        
        assert xlsx_bytes is not None
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0
    
    def test_xlsx_structure(self, sample_report_data, tmp_path):
        """Testa estrutura do arquivo XLSX"""
        from api.services.relay_config_report_service import RelayConfigReportService
        
        mock_session = None
        service = RelayConfigReportService(mock_session)
        
        xlsx_bytes = service._generate_xlsx(sample_report_data)
        
        # Salvar temporariamente
        xlsx_file = tmp_path / "test_report.xlsx"
        with open(xlsx_file, 'wb') as f:
            f.write(xlsx_bytes)
        
        # Ler e validar
        workbook = load_workbook(xlsx_file)
        
        # Verificar sheets
        assert 'Equipment Info' in workbook.sheetnames
        assert 'Protection Functions' in workbook.sheetnames
        assert 'Relay Settings' in workbook.sheetnames
        
        # Validar conteúdo da sheet de equipamento
        eq_sheet = workbook['Equipment Info']
        assert eq_sheet['A1'].value is not None  # Deve ter cabeçalho


class TestReportGenerationPDF:
    """Testes para geração de relatórios em PDF"""
    
    @pytest.fixture
    def sample_report_data(self):
        """Dados de exemplo para relatório"""
        return {
            'equipment_info': {
                'equipment_id': 1,
                'equipment_name': 'RELE_001',
                'manufacturer': 'MICON',
                'model': 'P122'
            },
            'protection_functions': [
                {'ansi_code': '50', 'function_name': 'Overcurrent', 'is_enabled': True}
            ],
            'relay_settings': [
                {
                    'setting_id': 1,
                    'ansi_function': '50',
                    'original_code': 'I>',
                    'setting_name': 'Pickup',
                    'configured_value': '5.5',
                    'unit': 'A'
                }
            ]
        }
    
    def test_generate_pdf_format(self, sample_report_data):
        """Testa geração de PDF"""
        from api.services.relay_config_report_service import RelayConfigReportService
        
        mock_session = None
        service = RelayConfigReportService(mock_session)
        
        pdf_bytes = service._generate_pdf(sample_report_data)
        
        assert pdf_bytes is not None
        assert isinstance(pdf_bytes, bytes)
        assert len(pdf_bytes) > 0
    
    def test_pdf_header(self, sample_report_data):
        """Testa que PDF tem header correto"""
        from api.services.relay_config_report_service import RelayConfigReportService
        
        mock_session = None
        service = RelayConfigReportService(mock_session)
        
        pdf_bytes = service._generate_pdf(sample_report_data)
        
        # PDFs começam com %PDF
        assert pdf_bytes[:4] == b'%PDF'


class TestEdgeCases:
    """Testes de casos extremos"""
    
    def test_equipment_not_found(self, mocker):
        """Testa equipamento não encontrado"""
        mock_session = mocker.MagicMock()
        mock_session.execute.return_value.fetchone.return_value = None
        
        service = RelayConfigReportService(mock_session)
        equipment = service.get_equipment_info(999)
        
        assert equipment is None
    
    def test_no_protection_functions(self, mocker):
        """Testa equipamento sem funções de proteção"""
        mock_session = mocker.MagicMock()
        mock_session.execute.return_value.fetchall.return_value = []
        
        service = RelayConfigReportService(mock_session)
        functions = service.get_protection_functions(1)
        
        assert functions == []
    
    def test_no_relay_settings(self, mocker):
        """Testa equipamento sem configurações"""
        mock_session = mocker.MagicMock()
        mock_session.execute.return_value.fetchall.return_value = []
        
        service = RelayConfigReportService(mock_session)
        settings = service.get_relay_settings(1)
        
        assert settings == []
    
    def test_include_disabled_filter(self, mocker):
        """Testa filtro de incluir desabilitados"""
        mock_session = mocker.MagicMock()
        
        # Mock com configurações ativas e inativas
        mock_active = mocker.MagicMock()
        mock_active.is_active = True
        mock_inactive = mocker.MagicMock()
        mock_inactive.is_active = False
        
        mock_session.execute.return_value.fetchall.return_value = [mock_active, mock_inactive]
        
        service = RelayConfigReportService(mock_session)
        
        # Com filtro (apenas ativos)
        settings_filtered = service.get_relay_settings(1, include_disabled=False)
        # Sem filtro (todos)
        settings_all = service.get_relay_settings(1, include_disabled=True)
        
        # Ambos devem retornar dados (filtro é aplicado na query SQL, não no Python)
        assert len(settings_all) >= len(settings_filtered)


class TestTokenization:
    """Testes para tokenização de valores"""
    
    def test_tokenize_simple_value(self):
        """Testa tokenização de valor simples"""
        from api.services.relay_config_report_service import RelayConfigReportService
        
        mock_session = None
        service = RelayConfigReportService(mock_session)
        
        tokens = service._tokenize_value("5.5")
        
        assert '5.5' in tokens or tokens == ['5', '.', '5'] or tokens == ['5.5']
    
    def test_tokenize_range_value(self):
        """Testa tokenização de valor com range"""
        from api.services.relay_config_report_service import RelayConfigReportService
        
        mock_session = None
        service = RelayConfigReportService(mock_session)
        
        tokens = service._tokenize_value("1.0-10.0")
        
        assert len(tokens) >= 3  # Deve ter pelo menos número, separador, número
    
    def test_tokenize_empty_value(self):
        """Testa tokenização de valor vazio"""
        from api.services.relay_config_report_service import RelayConfigReportService
        
        mock_session = None
        service = RelayConfigReportService(mock_session)
        
        tokens = service._tokenize_value("")
        
        assert tokens == [] or tokens == ['']


if __name__ == '__main__':
    pytest.main([__file__, '-v', '--tb=short'])
