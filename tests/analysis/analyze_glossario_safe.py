#!/usr/bin/env python3
"""
Análise SEGURA do glossário - sem travar o sistema
Lê apenas metadados básicos primeiro
"""

import openpyxl
from pathlib import Path

def analyze_glossario_metadata():
    """Análise leve - apenas estrutura"""
    glossario_path = Path('inputs/glossario/Dados_Glossario_Micon_Sepam.xlsx')
    
    if not glossario_path.exists():
        print(f'❌ Glossário não encontrado: {glossario_path}')
        return
    
    print('📊 ANÁLISE SEGURA DO GLOSSÁRIO')
    print('=' * 80)
    print(f'📁 Arquivo: {glossario_path.name}')
    print(f'💾 Tamanho: {glossario_path.stat().st_size / 1024:.1f} KB\n')
    
    # Abrir em modo read-only e data-only para evitar sobrecarga
    wb = openpyxl.load_workbook(glossario_path, read_only=True, data_only=True)
    
    print(f'📄 Total de abas: {len(wb.sheetnames)}\n')
    print('=' * 80)
    print('LISTA DE ABAS:')
    print('=' * 80)
    
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f'{i:2d}. {sheet_name:<40} ({ws.max_row:5d} linhas)')
    
    wb.close()
    print('\n✅ Análise de metadados concluída')

if __name__ == '__main__':
    analyze_glossario_metadata()
